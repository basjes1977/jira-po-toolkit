import requests
import os
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import time as _time

# Load Jira credentials from .jira_environment

def load_jira_env():
    env_path = Path(__file__).parent / ".jira_environment"
    if env_path.exists():
        env = {}
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("export "):
                    line = line[len("export "):]
                    if "=" in line:
                        k, v = line.split("=", 1)
                        env[k.strip()] = v.strip().strip('"')
        return env
    return {}

JIRA_ENV = load_jira_env()
JIRA_URL = JIRA_ENV.get("JT_JIRA_URL", "https://equinixjira.atlassian.net/").rstrip("/")
JIRA_EMAIL = JIRA_ENV.get("JT_JIRA_USERNAME")
JIRA_API_TOKEN = JIRA_ENV.get("JT_JIRA_PASSWORD")
BOARD_ID = JIRA_ENV.get("JT_JIRA_BOARD")


def get_sprints(state="closed", max_results=10):
    url = f"{JIRA_URL}/rest/agile/1.0/board/{BOARD_ID}/sprint?state={state}"
    resp = requests.get(url, auth=(JIRA_EMAIL, JIRA_API_TOKEN))
    resp.raise_for_status()
    sprints = resp.json().get("values", [])
    # Sort by endDate descending if available
    sprints = [s for s in sprints if s.get("endDate")]
    sprints.sort(key=lambda s: s["endDate"], reverse=True)
    return sprints[:max_results]


def get_issues(sprint_id):
    url = f"{JIRA_URL}/rest/agile/1.0/sprint/{sprint_id}/issue"
    issues = []
    start_at = 0
    while True:
        params = {"startAt": start_at, "maxResults": 50}
        resp = requests.get(url, params=params, auth=(JIRA_EMAIL, JIRA_API_TOKEN))
        resp.raise_for_status()
        data = resp.json()
        issues.extend(data["issues"])
        if start_at + 50 >= data["total"]:
            break
        start_at += 50
    return issues


def get_team_members(issues):
    members = set()
    for issue in issues:
        assignee = issue["fields"].get("assignee")
        if assignee and isinstance(assignee, dict):
            members.add(assignee.get("displayName", "Unknown"))
    return sorted(members)


def achieved_points_and_time(issues):
    points = 0
    time_logged = 0
    for issue in issues:
        fields = issue["fields"]
        status = fields.get("status", {}).get("name", "").lower()
        if status in ("done", "closed", "resolved"):
            story_points = fields.get("customfield_10024")
            if story_points not in (None, "?") and str(story_points).strip() != "":
                try:
                    points += float(story_points)
                except Exception:
                    pass
            if fields.get("timetracking") and isinstance(fields["timetracking"], dict):
                time_logged_val = fields["timetracking"].get("timeSpentSeconds")
                if time_logged_val not in (None, "", "?"):
                    try:
                        time_logged += int(time_logged_val)
                    except Exception:
                        pass
    return points, time_logged


def prompt_availability(members):
    print("\nEnter the number of days each team member is available in the coming sprint:")
    avail = {}
    for m in members:
        while True:
            try:
                days = float(input(f"  {m}: "))
                avail[m] = days
                break
            except Exception:
                print("  Please enter a number.")
    return avail

def try_save_workbook(wb, excel_name):
    while True:
        try:
            wb.save(excel_name)
            break
        except PermissionError:
            print(f"\nERROR: The file '{excel_name}' is open in Excel or another program. Please close it and press Enter to continue...")
            input()
            # Wait a moment before retrying
            _time.sleep(1)

def main():
    print("Fetching last 10 completed sprints...")
    sprints = get_sprints(state="closed", max_results=10)
    if not sprints:
        print("No completed sprints found.")
        return
    print("Sprints:")
    for s in sprints:
        print(f"  {s['name']} ({s.get('startDate', '')[:10]} to {s.get('endDate', '')[:10]})")
    results = []
    all_members = set()
    for s in sprints:
        issues = get_issues(s["id"])
        pts, tlog = achieved_points_and_time(issues)
        members = get_team_members(issues)
        all_members.update(members)
        results.append({"sprint": s, "points": pts, "time": tlog, "members": members})
    # Prompt for availability
    all_members = sorted(all_members)
    avail = prompt_availability(all_members)
    total_avail = sum(avail.values())
    # Estimate average available days in past sprints
    avg_avail_1 = len(results[0]["members"])*10 if results[0]["members"] else total_avail  # fallback: 10d per member
    avg_avail_3 = sum(len(r["members"])*10 for r in results[:3]) / 3 if results[:3] else total_avail
    avg_avail_5 = sum(len(r["members"])*10 for r in results[:5]) / 5 if results[:5] else total_avail
    avg_avail_10 = sum(len(r["members"])*10 for r in results[:10]) / 10 if results[:10] else total_avail
    # Calculate averages
    def fmt_time(sec):
        h = int(sec)//3600
        m = (int(sec)%3600)//60
        return f"{h}h {m}m"
    avg_pts_1 = results[0]["points"]
    avg_pts_3 = sum(r["points"] for r in results[:3]) / 3
    avg_pts_5 = sum(r["points"] for r in results[:5]) / 5
    avg_pts_10 = sum(r["points"] for r in results[:10]) / min(10, len(results))
    avg_time_1 = results[0]["time"]
    avg_time_3 = sum(r["time"] for r in results[:3]) / 3
    avg_time_5 = sum(r["time"] for r in results[:5]) / 5
    avg_time_10 = sum(r["time"] for r in results[:10]) / min(10, len(results))
    # Scale by availability
    scale_1 = total_avail / avg_avail_1 if avg_avail_1 else 1
    scale_3 = total_avail / avg_avail_3 if avg_avail_3 else 1
    scale_5 = total_avail / avg_avail_5 if avg_avail_5 else 1
    scale_10 = total_avail / avg_avail_10 if avg_avail_10 else 1
    print("\n--- Sprint Forecast ---")
    print(f"Team total available days (next sprint): {total_avail}")
    print("\n| Window   | Avg Points | Forecast Points| Avg Time | Forecast Time |")
    print("|----------|------------|----------------|----------|---------------|")
    print(f"| Last 1   | {avg_pts_1:.1f}        | {avg_pts_1*scale_1:.1f}            | {fmt_time(avg_time_1)}    | {fmt_time(avg_time_1*scale_1)}   |")
    print(f"| Last 3   | {avg_pts_3:.1f}       | {avg_pts_3*scale_3:.1f}           | {fmt_time(avg_time_3)}   | {fmt_time(avg_time_3*scale_3)}   |")
    print(f"| Last 5   | {avg_pts_5:.1f}        | {avg_pts_5*scale_5:.1f}           | {fmt_time(avg_time_5)}    | {fmt_time(avg_time_5*scale_5)}   |")
    print(f"| Last 10  | {avg_pts_10:.1f}       | {avg_pts_10*scale_10:.1f}           | {fmt_time(avg_time_10)}   | {fmt_time(avg_time_10*scale_10)}   |")
    print("\nNotes:")
    print("- Forecast is scaled by the ratio of available days (next sprint vs. past sprints, assuming 10d per member per sprint).")
    print("- You can adjust the 10d-per-sprint-per-member assumption in the script if your sprints are longer/shorter.")
    print("- Forecast is based on achieved (Done/Closed/Resolved) issues only.")

    # Excel export
    excel_name = "sprint_forecast_history.xlsx"
    # If file exists, append only new sprints to history
    if os.path.exists(excel_name):
        wb = openpyxl.load_workbook(excel_name)
        if "Sprint History" in wb.sheetnames:
            ws = wb["Sprint History"]
        else:
            ws = wb.create_sheet("Sprint History")
            ws.append(["Sprint Name", "Sprint Start", "Sprint End", "Achieved Story Points", "Achieved Time (h)"])
        # Get existing sprint names to avoid duplicates
        existing_sprints = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing_sprints.add(row[0])
        # Append only new sprints
        for r in results:
            s = r["sprint"]
            pts = r["points"]
            tlog = r["time"]
            start = s.get("startDate", "")[:10]
            end = s.get("endDate", "")[:10]
            if s["name"] not in existing_sprints:
                ws.append([s["name"], start, end, pts, round(tlog/3600, 2)])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sprint History"
        ws.append(["Sprint Name", "Sprint Start", "Sprint End", "Achieved Story Points", "Achieved Time (h)"])
        for r in results:
            s = r["sprint"]
            pts = r["points"]
            tlog = r["time"]
            start = s.get("startDate", "")[:10]
            end = s.get("endDate", "")[:10]
            ws.append([s["name"], start, end, pts, round(tlog/3600, 2)])
    # Add or update chart to Sprint History
    for obj in ws._charts:
        ws._charts.remove(obj)
    chart = LineChart()
    chart.title = "Achieved Story Points and Hours per Sprint"
    chart.y_axis.title = "Story Points / Hours (Done/Closed/Resolved)"
    chart.x_axis.title = "Sprint Name (chronological order)"
    data = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    chart.legend.position = 'b'
    chart.y_axis.majorGridlines = None
    chart.x_axis.tickLblPos = 'low'
    chart.x_axis.title = "Sprint Name (left=oldest, right=most recent)"
    chart.y_axis.title = "Achieved Story Points / Hours (sum for each sprint)"
    ws.add_chart(chart, f"G2")
    # Forecast sheet is always (re)created
    if "Forecast" in wb.sheetnames:
        del wb["Forecast"]
    ws2 = wb.create_sheet("Forecast")
    ws2.append(["Window", "Avg Points", "Forecast Points", "Avg Time (h)", "Forecast Time (h)"])
    ws2.append(["Last 1", round(avg_pts_1,1), round(avg_pts_1*scale_1,1), round(avg_time_1/3600,2), round(avg_time_1*scale_1/3600,2)])
    ws2.append(["Last 3", round(avg_pts_3,1), round(avg_pts_3*scale_3,1), round(avg_time_3/3600,2), round(avg_time_3*scale_3/3600,2)])
    ws2.append(["Last 5", round(avg_pts_5,1), round(avg_pts_5*scale_5,1), round(avg_time_5/3600,2), round(avg_time_5*scale_5/3600,2)])
    ws2.append(["Last 10", round(avg_pts_10,1), round(avg_pts_10*scale_10,1), round(avg_time_10/3600,2), round(avg_time_10*scale_10/3600,2)])
    # Add chart to Forecast
    chart2 = LineChart()
    chart2.title = "Forecasted Story Points and Hours (Next Sprint)"
    chart2.y_axis.title = "Forecasted Story Points / Hours"
    chart2.x_axis.title = "Window (Last N Sprints Used for Average)"
    data2 = Reference(ws2, min_col=2, max_col=5, min_row=1, max_row=ws2.max_row)
    cats2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 16
    chart2.height = 8
    chart2.legend.position = 'b'
    chart2.y_axis.majorGridlines = None
    chart2.x_axis.tickLblPos = 'low'
    chart2.x_axis.title = "Window (Last N Sprints Used for Average)"
    chart2.y_axis.title = "Forecasted Story Points / Hours (scaled for next sprint)"
    ws2.add_chart(chart2, f"G2")
    # Autosize columns
    for wsx in [ws, ws2]:
        for col in wsx.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            wsx.column_dimensions[col_letter].width = max_length + 2
    try_save_workbook(wb, excel_name)
    print(f"\nExcel file with sprint history and forecast saved as: {excel_name}")

if __name__ == "__main__":
    main()

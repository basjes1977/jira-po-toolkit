"""
Core sprint forecast logic for JiraPresentationTool.

This module provides the business logic for generating sprint forecasts.
It can be used by both the CLI and Flask web UI.
"""

import os
import logging
import time as _time
from pathlib import Path
from typing import Dict, List, Any, Optional, Callable, Tuple

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from jira_config import load_jira_env
from jira_metrics import achieved_points_and_time, get_recent_sprints, get_sprint_issues

# Set up logging
logger = logging.getLogger(__name__)

# Load Jira environment at module level for performance
JIRA_ENV = load_jira_env()
JIRA_URL = JIRA_ENV.get("JT_JIRA_URL", "https://equinixjira.atlassian.net/").rstrip("/")
JIRA_EMAIL = JIRA_ENV.get("JT_JIRA_USERNAME")
JIRA_API_TOKEN = JIRA_ENV.get("JT_JIRA_PASSWORD")
BOARD_ID = JIRA_ENV.get("JT_JIRA_BOARD")
FIELD_STORY_POINTS = JIRA_ENV.get("JT_JIRA_FIELD_STORY_POINTS", "customfield_10024")
AUTH = (JIRA_EMAIL, JIRA_API_TOKEN)


def get_team_members(issues: List[Dict]) -> List[str]:
    """Extract unique team member names from issues.

    Args:
        issues: List of Jira issue dicts

    Returns:
        Sorted list of team member display names
    """
    members = set()
    for issue in issues:
        assignee = issue["fields"].get("assignee")
        if assignee and isinstance(assignee, dict):
            members.add(assignee.get("displayName", "Unknown"))
    return sorted(members)


def try_save_workbook(wb, excel_path: Path):
    """Save Excel workbook with retry logic for file locks.

    Args:
        wb: openpyxl Workbook instance
        excel_path: Path to save the workbook

    Raises:
        Exception: If user cancels the retry loop
    """
    while True:
        try:
            wb.save(excel_path)
            break
        except PermissionError:
            logger.error(f"File '{excel_path}' is locked. Please close it.")
            # In non-interactive mode, raise the error
            raise


def get_sprint_history(
    max_sprints: int = 10,
    progress_callback: Optional[Callable[[str], None]] = None
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Fetch sprint history and extract team members.

    Args:
        max_sprints: Maximum number of completed sprints to fetch
        progress_callback: Optional callback for progress updates

    Returns:
        Tuple of (sprint_results, all_team_members):
            - sprint_results: List of dicts with sprint data
            - all_team_members: Sorted list of all team member names
    """
    def update_progress(msg: str):
        if progress_callback:
            progress_callback(msg)
        logger.info(msg)

    update_progress(f"Fetching last {max_sprints} completed sprints...")
    sprints = get_recent_sprints(JIRA_URL, BOARD_ID, AUTH, state="closed", max_results=max_sprints)

    if not sprints:
        logger.warning("No completed sprints found")
        return [], []

    results = []
    all_members = set()

    for idx, s in enumerate(sprints, 1):
        update_progress(f"Analyzing sprint {idx}/{len(sprints)}: {s['name']}...")
        issues = get_sprint_issues(JIRA_URL, s["id"], AUTH)
        pts, tlog = achieved_points_and_time(issues, FIELD_STORY_POINTS)
        members = get_team_members(issues)
        all_members.update(members)
        results.append({
            "sprint": s,
            "points": pts,
            "time": tlog,
            "members": members
        })

    update_progress(f"Fetched {len(results)} sprints with {len(all_members)} team members")
    return results, sorted(all_members)


def generate_sprint_forecast(
    team_availability: Dict[str, float],
    output_dir: Optional[Path] = None,
    max_sprints: int = 10,
    progress_callback: Optional[Callable[[str], None]] = None
) -> Dict[str, Any]:
    """Generate sprint forecast based on historical data and team availability.

    Args:
        team_availability: Dict mapping team member names to available days
        output_dir: Output directory for Excel file (None = current directory)
        max_sprints: Maximum number of completed sprints to analyze
        progress_callback: Optional callback for progress updates

    Returns:
        dict: {
            'success': bool,
            'file_path': Path or None,
            'forecast_windows': Dict with forecast data,
            'sprint_count': int,
            'error': str or None
        }
    """
    def update_progress(msg: str):
        if progress_callback:
            progress_callback(msg)
        logger.info(msg)

    try:
        # Set output directory
        if output_dir is None:
            output_dir = Path.cwd()
        else:
            output_dir = Path(output_dir)

        # Fetch sprint history
        results, all_members = get_sprint_history(max_sprints, progress_callback)

        if not results:
            return {
                'success': False,
                'file_path': None,
                'forecast_windows': {},
                'sprint_count': 0,
                'error': 'No completed sprints found'
            }

        # Calculate total availability for next sprint
        total_avail = sum(team_availability.values())

        # Estimate average available days in past sprints (assuming 10 days per member per sprint)
        avg_avail_1 = len(results[0]["members"]) * 10 if results[0]["members"] else total_avail
        avg_avail_3 = sum(len(r["members"]) * 10 for r in results[:3]) / 3 if len(results) >= 3 else avg_avail_1
        avg_avail_5 = sum(len(r["members"]) * 10 for r in results[:5]) / 5 if len(results) >= 5 else avg_avail_3
        avg_avail_10 = sum(len(r["members"]) * 10 for r in results[:10]) / min(10, len(results)) if results else total_avail

        # Calculate averages for each window
        def fmt_time(sec):
            h = int(sec) // 3600
            m = (int(sec) % 3600) // 60
            return f"{h}h {m}m"

        avg_pts_1 = results[0]["points"]
        avg_pts_3 = sum(r["points"] for r in results[:3]) / min(3, len(results))
        avg_pts_5 = sum(r["points"] for r in results[:5]) / min(5, len(results))
        avg_pts_10 = sum(r["points"] for r in results[:10]) / min(10, len(results))

        avg_time_1 = results[0]["time"]
        avg_time_3 = sum(r["time"] for r in results[:3]) / min(3, len(results))
        avg_time_5 = sum(r["time"] for r in results[:5]) / min(5, len(results))
        avg_time_10 = sum(r["time"] for r in results[:10]) / min(10, len(results))

        # Scale by availability ratio
        scale_1 = total_avail / avg_avail_1 if avg_avail_1 else 1
        scale_3 = total_avail / avg_avail_3 if avg_avail_3 else 1
        scale_5 = total_avail / avg_avail_5 if avg_avail_5 else 1
        scale_10 = total_avail / avg_avail_10 if avg_avail_10 else 1

        # Build forecast windows
        forecast_windows = {
            'last_1': {
                'avg_points': round(avg_pts_1, 1),
                'forecast_points': round(avg_pts_1 * scale_1, 1),
                'avg_time': round(avg_time_1 / 3600, 2),
                'forecast_time': round(avg_time_1 * scale_1 / 3600, 2),
                'avg_time_formatted': fmt_time(avg_time_1),
                'forecast_time_formatted': fmt_time(avg_time_1 * scale_1)
            },
            'last_3': {
                'avg_points': round(avg_pts_3, 1),
                'forecast_points': round(avg_pts_3 * scale_3, 1),
                'avg_time': round(avg_time_3 / 3600, 2),
                'forecast_time': round(avg_time_3 * scale_3 / 3600, 2),
                'avg_time_formatted': fmt_time(avg_time_3),
                'forecast_time_formatted': fmt_time(avg_time_3 * scale_3)
            },
            'last_5': {
                'avg_points': round(avg_pts_5, 1),
                'forecast_points': round(avg_pts_5 * scale_5, 1),
                'avg_time': round(avg_time_5 / 3600, 2),
                'forecast_time': round(avg_time_5 * scale_5 / 3600, 2),
                'avg_time_formatted': fmt_time(avg_time_5),
                'forecast_time_formatted': fmt_time(avg_time_5 * scale_5)
            },
            'last_10': {
                'avg_points': round(avg_pts_10, 1),
                'forecast_points': round(avg_pts_10 * scale_10, 1),
                'avg_time': round(avg_time_10 / 3600, 2),
                'forecast_time': round(avg_time_10 * scale_10 / 3600, 2),
                'avg_time_formatted': fmt_time(avg_time_10),
                'forecast_time_formatted': fmt_time(avg_time_10 * scale_10)
            },
            'total_availability': total_avail
        }

        # Generate Excel file
        update_progress("Generating Excel file...")
        excel_name = "sprint_forecast_history.xlsx"
        excel_path = output_dir / excel_name

        # Load or create workbook
        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
            if "Sprint History" in wb.sheetnames:
                ws = wb["Sprint History"]
            else:
                ws = wb.create_sheet("Sprint History")
                ws.append(["Sprint Name", "Sprint Start", "Sprint End", "Achieved Story Points", "Achieved Time (h)"])

            # Get existing sprint names to avoid duplicates
            existing_sprints = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    existing_sprints.add(row[0])

            # Append only new sprints
            for r in results:
                s = r["sprint"]
                if s["name"] not in existing_sprints:
                    start = s.get("startDate", "")[:10]
                    end = s.get("endDate", "")[:10]
                    ws.append([s["name"], start, end, r["points"], round(r["time"] / 3600, 2)])
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sprint History"
            ws.append(["Sprint Name", "Sprint Start", "Sprint End", "Achieved Story Points", "Achieved Time (h)"])
            for r in results:
                s = r["sprint"]
                start = s.get("startDate", "")[:10]
                end = s.get("endDate", "")[:10]
                ws.append([s["name"], start, end, r["points"], round(r["time"] / 3600, 2)])

        # Add/update chart to Sprint History
        for obj in ws._charts:
            ws._charts.remove(obj)

        chart = LineChart()
        chart.title = "Achieved Story Points and Hours per Sprint"
        chart.y_axis.title = "Story Points / Hours (Done/Closed/Resolved)"
        chart.x_axis.title = "Sprint Name (chronological order)"
        data = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        chart.height = 10
        chart.legend.position = 'b'
        chart.y_axis.majorGridlines = None
        chart.x_axis.tickLblPos = 'low'
        chart.x_axis.title = "Sprint Name (left=oldest, right=most recent)"
        chart.y_axis.title = "Achieved Story Points / Hours (sum for each sprint)"
        ws.add_chart(chart, "G2")

        # Forecast sheet (always recreated)
        if "Forecast" in wb.sheetnames:
            del wb["Forecast"]
        ws2 = wb.create_sheet("Forecast")
        ws2.append(["Window", "Avg Points", "Forecast Points", "Avg Time (h)", "Forecast Time (h)"])
        ws2.append(["Last 1",
                    forecast_windows['last_1']['avg_points'],
                    forecast_windows['last_1']['forecast_points'],
                    forecast_windows['last_1']['avg_time'],
                    forecast_windows['last_1']['forecast_time']])
        ws2.append(["Last 3",
                    forecast_windows['last_3']['avg_points'],
                    forecast_windows['last_3']['forecast_points'],
                    forecast_windows['last_3']['avg_time'],
                    forecast_windows['last_3']['forecast_time']])
        ws2.append(["Last 5",
                    forecast_windows['last_5']['avg_points'],
                    forecast_windows['last_5']['forecast_points'],
                    forecast_windows['last_5']['avg_time'],
                    forecast_windows['last_5']['forecast_time']])
        ws2.append(["Last 10",
                    forecast_windows['last_10']['avg_points'],
                    forecast_windows['last_10']['forecast_points'],
                    forecast_windows['last_10']['avg_time'],
                    forecast_windows['last_10']['forecast_time']])

        # Add chart to Forecast
        chart2 = LineChart()
        chart2.title = "Forecasted Story Points and Hours (Next Sprint)"
        chart2.y_axis.title = "Forecasted Story Points / Hours"
        chart2.x_axis.title = "Window (Last N Sprints Used for Average)"
        data2 = Reference(ws2, min_col=2, max_col=5, min_row=1, max_row=ws2.max_row)
        cats2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.width = 16
        chart2.height = 8
        chart2.legend.position = 'b'
        chart2.y_axis.majorGridlines = None
        chart2.x_axis.tickLblPos = 'low'
        chart2.x_axis.title = "Window (Last N Sprints Used for Average)"
        chart2.y_axis.title = "Forecasted Story Points / Hours (scaled for next sprint)"
        ws2.add_chart(chart2, "G2")

        # Autosize columns
        for wsx in [ws, ws2]:
            for col in wsx.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                wsx.column_dimensions[col_letter].width = max_length + 2

        # Save workbook
        update_progress("Saving Excel file...")
        try_save_workbook(wb, excel_path)

        update_progress("Done!")

        return {
            'success': True,
            'file_path': excel_path,
            'forecast_windows': forecast_windows,
            'sprint_count': len(results),
            'error': None
        }

    except Exception as e:
        logger.exception("Error generating sprint forecast")
        return {
            'success': False,
            'file_path': None,
            'forecast_windows': {},
            'sprint_count': 0,
            'error': str(e)
        }
